#!/usr/bin/env python3
"""
Build a normalized SQLite database from SETSmart REIT/PFPO financial downloads.

Usage:
  python3 scripts/build_financial_database.py

The script scans *_FS_*.xlsx files in the project root, rebuilds
data/financials.sqlite, and exports data/financials_dashboard_data.js for the
HTML dashboard.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sqlite3
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DB_PATH = ROOT / "data" / "financials.sqlite"
DASHBOARD_DATA_PATH = ROOT / "data" / "financials_dashboard_data.js"
HEADER_RE = re.compile(
    r"Q(?P<quarter>[1-4])\s*/\s*(?P<year>\d{4})\s*"
    r"\((?P<start>\d{2}/\d{2}/\d{2})\s*-\s*(?P<end>\d{2}/\d{2}/\d{2})\)\s*"
    r"(?P<entity_type>.+)"
)


def norm_text(value: Any) -> str:
    return " ".join(str(value or "").strip().split())


def item_key(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")


def item_group(name: str, key: str) -> str:
    text = f"{name} {key}".lower()
    if "net_investment_income" in text or "increase_decrease" in text or "comprehensive_income" in text:
        return "Profit / Comprehensive Income"
    if "expense" in text or "fee" in text or "cost" in text or "tax" in text or "amortisation" in text:
        return "Expenses"
    if "revenue" in text or "income" in text:
        return "Revenue"
    if "gain" in text or "loss" in text or "fair_value" in text or "foreign_currency" in text:
        return "Gains / Losses"
    return "Other"


def parse_dmy(value: str) -> str:
    return dt.datetime.strptime(value, "%d/%m/%y").date().isoformat()


def parse_period_header(value: Any) -> dict[str, Any] | None:
    text = norm_text(value)
    match = HEADER_RE.match(text)
    if not match:
        return None
    reported_quarter = int(match.group("quarter"))
    reported_year = int(match.group("year"))
    start_date = dt.datetime.strptime(match.group("start"), "%d/%m/%y").date()
    end_date = dt.datetime.strptime(match.group("end"), "%d/%m/%y").date()
    calendar_quarter = ((end_date.month - 1) // 3) + 1
    calendar_year = end_date.year
    return {
        "period_label": f"Q{calendar_quarter} {calendar_year}",
        "reported_label": f"Q{reported_quarter} {reported_year}",
        "fiscal_year": calendar_year,
        "fiscal_quarter": calendar_quarter,
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "entity_type": norm_text(match.group("entity_type")),
        "sort_key": calendar_year * 10 + calendar_quarter,
    }


def as_float(value: Any) -> float | None:
    if value in (None, "", "-", "N/A"):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text or text == "-":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def connect() -> sqlite3.Connection:
    DB_PATH.parent.mkdir(exist_ok=True)
    if DB_PATH.exists():
        DB_PATH.unlink()
    con = sqlite3.connect(DB_PATH)
    con.execute("PRAGMA foreign_keys = ON")
    con.executescript(
        """
        CREATE TABLE import_runs (
            id INTEGER PRIMARY KEY,
            imported_at TEXT NOT NULL,
            source_directory TEXT NOT NULL,
            file_count INTEGER NOT NULL
        );

        CREATE TABLE source_files (
            id INTEGER PRIMARY KEY,
            import_run_id INTEGER NOT NULL REFERENCES import_runs(id),
            file_name TEXT NOT NULL UNIQUE,
            file_path TEXT NOT NULL,
            file_mtime TEXT NOT NULL,
            export_timestamp TEXT
        );

        CREATE TABLE entities (
            symbol TEXT PRIMARY KEY,
            entity_name TEXT,
            market TEXT,
            sector TEXT,
            industry TEXT,
            latest_price REAL,
            last_source_file_id INTEGER REFERENCES source_files(id)
        );

        CREATE TABLE periods (
            id INTEGER PRIMARY KEY,
            period_label TEXT NOT NULL,
            reported_label TEXT NOT NULL,
            fiscal_year INTEGER NOT NULL,
            fiscal_quarter INTEGER NOT NULL,
            start_date TEXT NOT NULL,
            end_date TEXT NOT NULL,
            entity_type TEXT NOT NULL,
            sort_key INTEGER NOT NULL,
            UNIQUE(start_date, end_date)
        );

        CREATE TABLE financial_items (
            id INTEGER PRIMARY KEY,
            statement_type TEXT NOT NULL,
            item_name TEXT NOT NULL,
            item_key TEXT NOT NULL,
            indent_level INTEGER NOT NULL,
            parent_item_id INTEGER REFERENCES financial_items(id),
            display_order INTEGER NOT NULL,
            UNIQUE(statement_type, item_key, indent_level)
        );

        CREATE TABLE financial_facts (
            id INTEGER PRIMARY KEY,
            symbol TEXT NOT NULL REFERENCES entities(symbol),
            period_id INTEGER NOT NULL REFERENCES periods(id),
            item_id INTEGER NOT NULL REFERENCES financial_items(id),
            value_kbaht REAL,
            source_file_id INTEGER NOT NULL REFERENCES source_files(id),
            UNIQUE(symbol, period_id, item_id)
        );

        CREATE VIEW v_financial_facts AS
        SELECT
            f.symbol,
            e.entity_name,
            p.period_label,
            p.fiscal_year,
            p.fiscal_quarter,
            p.start_date,
            p.end_date,
            i.statement_type,
            i.item_name,
            i.item_key,
            i.indent_level,
            f.value_kbaht,
            sf.file_name AS source_file
        FROM financial_facts f
        JOIN entities e ON e.symbol = f.symbol
        JOIN periods p ON p.id = f.period_id
        JOIN financial_items i ON i.id = f.item_id
        JOIN source_files sf ON sf.id = f.source_file_id;

        CREATE VIEW v_period_changes AS
        SELECT
            symbol,
            item_name,
            item_key,
            period_label,
            fiscal_year,
            fiscal_quarter,
            value_kbaht,
            value_kbaht - LAG(value_kbaht) OVER (
                PARTITION BY symbol, item_key ORDER BY fiscal_year, fiscal_quarter
            ) AS qoq_change_kbaht,
            CASE
                WHEN LAG(value_kbaht) OVER (
                    PARTITION BY symbol, item_key ORDER BY fiscal_year, fiscal_quarter
                ) IS NOT NULL
                AND LAG(value_kbaht) OVER (
                    PARTITION BY symbol, item_key ORDER BY fiscal_year, fiscal_quarter
                ) != 0
                THEN 100.0 * (value_kbaht - LAG(value_kbaht) OVER (
                    PARTITION BY symbol, item_key ORDER BY fiscal_year, fiscal_quarter
                )) / ABS(LAG(value_kbaht) OVER (
                    PARTITION BY symbol, item_key ORDER BY fiscal_year, fiscal_quarter
                ))
            END AS qoq_change_pct,
            value_kbaht - LAG(value_kbaht, 4) OVER (
                PARTITION BY symbol, item_key ORDER BY fiscal_year, fiscal_quarter
            ) AS yoy_change_kbaht,
            CASE
                WHEN LAG(value_kbaht, 4) OVER (
                    PARTITION BY symbol, item_key ORDER BY fiscal_year, fiscal_quarter
                ) IS NOT NULL
                AND LAG(value_kbaht, 4) OVER (
                    PARTITION BY symbol, item_key ORDER BY fiscal_year, fiscal_quarter
                ) != 0
                THEN 100.0 * (value_kbaht - LAG(value_kbaht, 4) OVER (
                    PARTITION BY symbol, item_key ORDER BY fiscal_year, fiscal_quarter
                )) / ABS(LAG(value_kbaht, 4) OVER (
                    PARTITION BY symbol, item_key ORDER BY fiscal_year, fiscal_quarter
                ))
            END AS yoy_change_pct
        FROM v_financial_facts;
        """
    )
    return con


def get_or_create_period(con: sqlite3.Connection, period: dict[str, Any]) -> int:
    con.execute(
        """
        INSERT OR IGNORE INTO periods
        (period_label, reported_label, fiscal_year, fiscal_quarter, start_date, end_date, entity_type, sort_key)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            period["period_label"],
            period["reported_label"],
            period["fiscal_year"],
            period["fiscal_quarter"],
            period["start_date"],
            period["end_date"],
            period["entity_type"],
            period["sort_key"],
        ),
    )
    return int(
        con.execute(
            "SELECT id FROM periods WHERE start_date = ? AND end_date = ?",
            (period["start_date"], period["end_date"]),
        ).fetchone()[0]
    )


def get_or_create_item(
    con: sqlite3.Connection,
    statement_type: str,
    name: str,
    indent_level: int,
    parent_id: int | None,
    display_order: int,
) -> int:
    key = item_key(name)
    con.execute(
        """
        INSERT OR IGNORE INTO financial_items
        (statement_type, item_name, item_key, indent_level, parent_item_id, display_order)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (statement_type, name, key, indent_level, parent_id, display_order),
    )
    row = con.execute(
        """
        SELECT id FROM financial_items
        WHERE statement_type = ? AND item_key = ? AND indent_level = ?
        """,
        (statement_type, key, indent_level),
    ).fetchone()
    return int(row[0])


def import_file(con: sqlite3.Connection, import_run_id: int, path: Path) -> int:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    symbol = norm_text(ws["B3"].value)
    statement_type = norm_text(ws["B4"].value)
    entity_name = norm_text(ws["A9"].value)
    market = norm_text(ws["B11"].value)
    sector = norm_text(ws["C11"].value)
    industry = norm_text(ws["D11"].value)
    latest_price = as_float(ws["B10"].value)
    export_timestamp = norm_text(ws["C1"].value)

    source_id = con.execute(
        """
        INSERT INTO source_files
        (import_run_id, file_name, file_path, file_mtime, export_timestamp)
        VALUES (?, ?, ?, ?, ?)
        """,
        (
            import_run_id,
            path.name,
            str(path),
            dt.datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds"),
            export_timestamp,
        ),
    ).lastrowid

    con.execute(
        """
        INSERT INTO entities
        (symbol, entity_name, market, sector, industry, latest_price, last_source_file_id)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(symbol) DO UPDATE SET
            entity_name = excluded.entity_name,
            market = excluded.market,
            sector = excluded.sector,
            industry = excluded.industry,
            latest_price = excluded.latest_price,
            last_source_file_id = excluded.last_source_file_id
        """,
        (symbol, entity_name, market, sector, industry, latest_price, source_id),
    )

    header_row = next(ws.iter_rows(min_row=19, max_row=19, values_only=True))
    period_cols: list[tuple[int, int]] = []
    for col_idx, raw_header in enumerate(header_row, start=1):
        period = parse_period_header(raw_header)
        if period:
            period_cols.append((col_idx, get_or_create_period(con, period)))

    parent_stack: dict[int, int] = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=20, max_row=ws.max_row, values_only=True), start=20):
        raw_item = row[0] if row else None
        if not isinstance(raw_item, str) or not raw_item.strip():
            continue
        name = norm_text(raw_item)
        if name.startswith("*") or name.startswith("Information on") or name.startswith("Restatement means"):
            continue
        values_by_period = [
            (period_id, as_float(row[col_idx - 1] if col_idx - 1 < len(row) else None))
            for col_idx, period_id in period_cols
        ]
        values_by_period = [(period_id, value) for period_id, value in values_by_period if value is not None]
        if not values_by_period:
            continue
        indent_level = (len(raw_item) - len(raw_item.lstrip(" "))) // 2
        parent_id = None
        lower_parents = [level for level in parent_stack if level < indent_level]
        if lower_parents:
            parent_id = parent_stack[max(lower_parents)]
        item_id = get_or_create_item(con, statement_type, name, indent_level, parent_id, row_idx)
        parent_stack[indent_level] = item_id
        for old_level in [level for level in parent_stack if level > indent_level]:
            parent_stack.pop(old_level, None)

        for period_id, value in values_by_period:
            con.execute(
                """
                INSERT INTO financial_facts (symbol, period_id, item_id, value_kbaht, source_file_id)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(symbol, period_id, item_id) DO UPDATE SET
                    value_kbaht = excluded.value_kbaht,
                    source_file_id = excluded.source_file_id
                """,
                (symbol, period_id, item_id, value, source_id),
            )
    wb.close()
    return 1


def export_dashboard_data(con: sqlite3.Connection) -> None:
    rows = con.execute(
        """
        SELECT symbol, entity_name, market, sector, industry, latest_price
        FROM entities
        ORDER BY symbol
        """
    ).fetchall()
    entities = [
        {
            "symbol": r[0],
            "name": r[1],
            "market": r[2],
            "sector": r[3],
            "industry": r[4],
            "latestPrice": r[5],
        }
        for r in rows
    ]
    periods = [
        {
            "id": r[0],
            "label": r[1],
            "year": r[2],
            "quarter": r[3],
            "startDate": r[4],
            "endDate": r[5],
            "sortKey": r[6],
        }
        for r in con.execute(
            """
            SELECT id, period_label, fiscal_year, fiscal_quarter, start_date, end_date, sort_key
            FROM periods
            ORDER BY sort_key
            """
        )
    ]
    items = [
        {
            "id": r[0],
            "name": r[1],
            "key": r[2],
            "indent": r[3],
            "order": r[4],
            "group": item_group(r[1], r[2]),
            "factCount": r[5],
        }
        for r in con.execute(
            """
            SELECT
                i.id,
                i.item_name,
                i.item_key,
                i.indent_level,
                MIN(i.display_order),
                COUNT(f.id) AS fact_count
            FROM financial_items i
            JOIN financial_facts f ON f.item_id = i.id
            GROUP BY i.item_key, i.item_name, i.indent_level
            HAVING fact_count > 0
            ORDER BY MIN(i.display_order), i.item_name
            """
        )
    ]
    facts = [
        {"symbol": r[0], "periodId": r[1], "itemKey": r[2], "value": r[3]}
        for r in con.execute(
            """
            SELECT f.symbol, f.period_id, i.item_key, f.value_kbaht
            FROM financial_facts f
            JOIN financial_items i ON i.id = f.item_id
            ORDER BY f.symbol, f.period_id, i.display_order
            """
        )
    ]
    latest = con.execute("SELECT MAX(imported_at), MAX(file_count) FROM import_runs").fetchone()
    payload = {
        "metadata": {
            "generatedAt": dt.datetime.now().isoformat(timespec="seconds"),
            "importedAt": latest[0],
            "fileCount": latest[1],
            "unit": "k.Baht",
            "database": str(DB_PATH.relative_to(ROOT)),
        },
        "entities": entities,
        "periods": periods,
        "items": items,
        "facts": facts,
    }
    DASHBOARD_DATA_PATH.write_text(
        "window.FINANCIAL_DATA = "
        + json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", type=Path, default=ROOT)
    parser.add_argument("--pattern", default="*_FS_*.xlsx")
    args = parser.parse_args()

    files = sorted(args.input_dir.glob(args.pattern))
    if not files:
        raise SystemExit(f"No files matched {args.pattern!r} in {args.input_dir}")

    con = connect()
    imported_at = dt.datetime.now().isoformat(timespec="seconds")
    run_id = con.execute(
        "INSERT INTO import_runs (imported_at, source_directory, file_count) VALUES (?, ?, ?)",
        (imported_at, str(args.input_dir), len(files)),
    ).lastrowid
    for path in files:
        import_file(con, int(run_id), path)
    con.commit()
    export_dashboard_data(con)

    entity_count = con.execute("SELECT COUNT(*) FROM entities").fetchone()[0]
    fact_count = con.execute("SELECT COUNT(*) FROM financial_facts").fetchone()[0]
    period_count = con.execute("SELECT COUNT(*) FROM periods").fetchone()[0]
    item_count = con.execute("SELECT COUNT(*) FROM financial_items").fetchone()[0]
    con.close()
    print(f"Imported {len(files)} files for {entity_count} symbols")
    print(f"Created {DB_PATH.relative_to(ROOT)}")
    print(f"Facts: {fact_count:,} | Periods: {period_count} | Items: {item_count}")
    print(f"Exported {DASHBOARD_DATA_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
